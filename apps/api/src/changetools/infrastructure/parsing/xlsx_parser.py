"""Xlsx parser using ``openpyxl``.

The Obsidian data pack uses named tables (KPITracker, RiskRegister, …). We
prefer those when present because they carry headers and ranges; sheets without
declared tables fall back to a heuristic header-row detection.
"""

from __future__ import annotations

from datetime import date, datetime
from typing import Any, BinaryIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from changetools.infrastructure.parsing.base import ParsedBlock, ParsedDocument


class XlsxParser:
    name = "xlsx"
    supported_extensions = (".xlsx",)

    def parse(self, stream: BinaryIO, *, filename: str) -> ParsedDocument:
        wb = load_workbook(stream, data_only=True, read_only=False)
        blocks: list[ParsedBlock] = []
        raw_lines: list[str] = []
        table_catalogue: list[dict[str, Any]] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_blocks, sheet_tables = self._parse_sheet(ws, sheet_name)
            blocks.extend(sheet_blocks)
            table_catalogue.extend(sheet_tables)
            for b in sheet_blocks:
                raw_lines.append(b.text)

        meta = {
            "filename": filename,
            "sheet_count": len(wb.sheetnames),
            "sheet_names": list(wb.sheetnames),
            "tables": table_catalogue,
        }
        return ParsedDocument(
            raw_text="\n".join(raw_lines),
            blocks=blocks,
            meta=meta,
        )

    def _parse_sheet(
        self, ws: Worksheet, sheet_name: str
    ) -> tuple[list[ParsedBlock], list[dict[str, Any]]]:
        blocks: list[ParsedBlock] = []
        catalogue: list[dict[str, Any]] = []

        tables = list(ws.tables.items()) if ws.tables else []
        if tables:
            for table_name, table_ref in tables:
                ref = table_ref if isinstance(table_ref, str) else table_ref.ref
                rows = list(ws[ref])
                if not rows:
                    continue
                headers = [self._cell_to_str(c.value) for c in rows[0]]
                blocks.append(
                    ParsedBlock(
                        kind="table_header",
                        text=f"[{sheet_name} · {table_name}] " + " | ".join(headers),
                        meta={
                            "sheet": sheet_name,
                            "table": table_name,
                            "headers": headers,
                        },
                    )
                )
                rows_added = 0
                for row_idx, row in enumerate(rows[1:], start=1):
                    cells = [self._cell_to_str(c.value) for c in row]
                    if not any(cells):
                        continue
                    cells_dict = dict(zip(headers, cells, strict=False))
                    text = self._format_row(sheet_name, table_name, headers, cells_dict)
                    blocks.append(
                        ParsedBlock(
                            kind="table_row",
                            text=text,
                            meta={
                                "sheet": sheet_name,
                                "table": table_name,
                                "row_index": row_idx,
                                "headers": headers,
                                "cells": cells_dict,
                            },
                        )
                    )
                    rows_added += 1
                catalogue.append(
                    {
                        "sheet": sheet_name,
                        "table": table_name,
                        "headers": headers,
                        "row_count": rows_added,
                    }
                )
        else:
            blocks.extend(self._parse_freeform_sheet(ws, sheet_name))
        return blocks, catalogue

    def _parse_freeform_sheet(self, ws: Worksheet, sheet_name: str) -> list[ParsedBlock]:
        """For sheets with no declared tables: capture each non-empty cell as a block."""
        blocks: list[ParsedBlock] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            row_text_parts: list[str] = []
            for cell in row:
                value = self._cell_to_str(cell.value)
                if value:
                    coord = f"{get_column_letter(cell.column)}{cell.row}"
                    row_text_parts.append(f"{coord}: {value}")
            if not row_text_parts:
                continue
            blocks.append(
                ParsedBlock(
                    kind="paragraph",
                    text=f"[{sheet_name}] " + "; ".join(row_text_parts),
                    meta={"sheet": sheet_name, "row_index": row_idx},
                )
            )
        return blocks

    @staticmethod
    def _cell_to_str(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, float):
            # Trim trailing zeros from floats — KPI percentages render cleaner.
            if value.is_integer():
                return str(int(value))
            return f"{value:.4g}"
        return str(value).strip()

    @staticmethod
    def _format_row(
        sheet: str,
        table: str,
        headers: list[str],
        cells: dict[str, str],
    ) -> str:
        prefix = f"[{sheet} · {table}]"
        body = " | ".join(f"{h}: {v}" for h, v in cells.items() if h and v)
        return f"{prefix} {body}".strip()
